import os
import json
import pandas as pd
import xlsxwriter
from flask import Flask, send_file, render_template, jsonify
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from io import BytesIO
from datetime import datetime

def flatten_json(json_obj, parent_key='', separator='_'):
    items = {}
    for key, value in json_obj.items():
        new_key = f"{parent_key}{separator}{key}" if parent_key else key
        if isinstance(value, dict):
            items.update(flatten_json(value, new_key, separator))
        else:
            items[new_key] = value
    return items

def convert_json_to_csv(json_filename, csv_filename):
    try:
        with open(json_filename, 'r') as json_file:
            json_data = json.load(json_file)
        
        rows = []
        for entry in json_data:
            flat_entry = flatten_json(entry)
            rows.append(flat_entry)
        
        df = pd.DataFrame(rows)
        df.to_csv(csv_filename, index=False)
        
        print(f"JSON data from '{json_filename}' successfully converted to CSV '{csv_filename}'.")
    except Exception as e:
        print("An error occurred:", str(e))

def csv_to_excel(csv_filename, excel_filename):
    try:
        # Read CSV data into a pandas DataFrame
        data = pd.read_csv(csv_filename)
        
        # Create a new Excel workbook and sheet
        wb = Workbook()
        sheet = wb.active
        sheet.title = "default"
        
        # Write headers and data to the Excel sheet
        headers = list(data.columns)
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row_num, row_data in data.iterrows():
            for col_num, cell_data in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num + 2, column=col_num, value=cell_data)
                cell.alignment = Alignment(vertical="center")
        
        # Save the Excel workbook
        wb.save(excel_filename)
        
        print(f"CSV data from '{csv_filename}' successfully converted to Excel '{excel_filename}'.")
    except Exception as e:
        print("An error occurred:", str(e))

def clear_files(directory):
    extensions_to_clear = ['.json', '.csv', '.xlsx']
    
    for root, dirs, files in os.walk(directory):
        for file in files:
            if any(file.endswith(ext) for ext in extensions_to_clear):
                file_path = os.path.join(root, file)
                os.remove(file_path)
                print(f"Removed: {file_path}")

